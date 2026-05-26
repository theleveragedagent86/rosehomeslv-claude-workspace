#!/usr/bin/env python3
"""
Listing Leads Content Scraper
==============================
Scrapes the Listing Attraction Plan from listingleads.com, downloads all
content/assets, organizes into folders by category, and generates a
professional Excel workbook.

Usage:
    python3 scrape_listing_leads.py --tab 1                    # Full scrape
    python3 scrape_listing_leads.py --tab 1 --discover         # Plan page DOM discovery
    python3 scrape_listing_leads.py --tab 1 --discover-detail  # Campaign detail page discovery
    python3 scrape_listing_leads.py --tab 1 --weeks 1          # Test with 1 week
    python3 scrape_listing_leads.py --tab 1 --resume           # Resume from checkpoint
    python3 scrape_listing_leads.py --tab 1 --no-download      # Skip file downloads
    python3 scrape_listing_leads.py --tab 1 --excel-only       # Regenerate Excel from JSON

Requirements: macOS, Google Chrome (with View > Developer > Allow JavaScript
from Apple Events enabled), logged into listingleads.com/plan.
"""

import sys
import os
import re
import json
import subprocess
import argparse
import time
import urllib.parse
from pathlib import Path
from datetime import datetime, timedelta

try:
    import requests
except ImportError:
    print("ERROR: requests not installed. Run: pip3 install requests")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)


# ─── Configuration ───────────────────────────────────────────────────────────

TAB = "tab 1"
OUTPUT_DIR = Path("/Users/ryanrose/Downloads/Claude/Listing Leads Content")
CHECKPOINT_FILE = Path("/tmp/listing_leads_progress.json")
TARGET_DATE = datetime(2025, 9, 1)  # Scrape back to week of Sep 1, 2025
PLAN_URL = "https://www.listingleads.com/plan"

# Excel color scheme by top-level category
CATEGORY_COLORS = {
    "Email Campaigns": "E8EDF3",          # light purple
    "Social Shareables": "E0F5F3",        # teal
    "Phone & Text Scripts": "F5F3E0",     # light yellow
    "Direct Mail Templates": "F5EDE4",    # warm peach
    "Video": "E4F5ED",                    # light green
    "Other": "F5F5F5",                    # light gray
}

# Map category labels to folder names
CATEGORY_FOLDERS = {
    "Email Campaigns": "Email Campaigns",
    "Social Shareables": "Social Shareables",
    "Phone & Text Scripts": "Phone & Text Scripts",
    "Direct Mail Templates": "Direct Mail Templates",
}


# ─── Chrome Control ──────────────────────────────────────────────────────────

def chrome_js(js_code):
    """Execute JS in Chrome tab via AppleScript using temp file."""
    tmp = Path("/tmp/ll-scrape-js-cmd.js")
    tmp.write_text(js_code)

    applescript = f'''
    set jsCode to read POSIX file "{str(tmp)}"
    tell application "Google Chrome"
        tell {TAB} of window 1
            execute javascript jsCode
        end tell
    end tell
    '''
    try:
        result = subprocess.run(
            ['osascript', '-e', applescript],
            capture_output=True, text=True, timeout=60
        )
        out = result.stdout.strip()
        if out == "missing value":
            return None
        return out
    except subprocess.TimeoutExpired:
        print("  WARNING: Chrome JS execution timed out")
        return None


def chrome_activate():
    subprocess.run(
        ['osascript', '-e', 'tell application "Google Chrome" to activate'],
        capture_output=True
    )
    time.sleep(0.3)


def wait_for_element(selector, timeout=10):
    """Wait for a DOM element to exist."""
    for _ in range(timeout * 2):
        result = chrome_js(f"!!document.querySelector('{selector}')")
        if result == "true":
            return True
        time.sleep(0.5)
    return False


def get_current_url():
    """Get the current URL of the Chrome tab."""
    applescript = f'''
    tell application "Google Chrome"
        tell {TAB} of window 1
            get URL
        end tell
    end tell
    '''
    result = subprocess.run(['osascript', '-e', applescript], capture_output=True, text=True)
    return result.stdout.strip()


def navigate_to(url):
    """Navigate Chrome tab to a URL."""
    applescript = f'''
    tell application "Google Chrome"
        tell {TAB} of window 1
            set URL to "{url}"
        end tell
    end tell
    '''
    subprocess.run(['osascript', '-e', applescript], capture_output=True, text=True)
    time.sleep(3)


def sanitize_filename(name, max_len=80):
    """Create a safe filename from text."""
    name = re.sub(r'[<>:"/\\|?*\[\]]', '', name)
    name = re.sub(r'\s+', '_', name.strip())
    name = re.sub(r'_+', '_', name)
    if len(name) > max_len:
        name = name[:max_len]
    return name or "untitled"


# ─── DOM Discovery ───────────────────────────────────────────────────────────

def discover_detail_page():
    """Discover the structure of a campaign detail page."""
    print("\n" + "=" * 70)
    print("CAMPAIGN DETAIL PAGE DISCOVERY")
    print("=" * 70)

    url = get_current_url()
    print(f"Current URL: {url}")

    if "/campaigns/" not in url:
        print("\nWARNING: Navigate to a campaign detail page first!")
        print("e.g. click on any campaign card from the plan page")
        return

    print("\n--- FULL PAGE TEXT ---")
    result = chrome_js("""
    (function() {
        var main = document.querySelector('main') || document.body;
        return main.innerText.substring(0, 10000);
    })()
    """)
    print(result or "(no results)")

    print("\n--- CONTENT SECTIONS ---")
    result = chrome_js("""
    (function() {
        var els = document.querySelectorAll(
            '[class*="content"], [class*="body"], [class*="description"], ' +
            '[class*="script"], [class*="copy"], [class*="template"], ' +
            'article, section, [class*="campaign"], [class*="detail"]'
        );
        var items = [];
        els.forEach(function(el) {
            var rect = el.getBoundingClientRect();
            if (rect.width > 100 && rect.height > 50 && items.length < 30) {
                items.push({
                    tag: el.tagName,
                    classes: el.className.toString().substring(0, 150),
                    textPreview: el.textContent.trim().substring(0, 300),
                    childCount: el.children.length,
                    id: el.id
                });
            }
        });
        return JSON.stringify(items, null, 2);
    })()
    """)
    print(result or "(no results)")

    print("\n--- DOWNLOAD LINKS & BUTTONS ---")
    result = chrome_js("""
    (function() {
        var els = document.querySelectorAll(
            'a[href*="download"], a[download], a[href*=".pdf"], a[href*=".png"], ' +
            'a[href*=".jpg"], a[href*=".zip"], a[href*=".docx"], ' +
            'a[href*="canva"], a[href*="drive.google"], a[href*="dropbox"], ' +
            'button[class*="download"], [class*="download"], ' +
            'a[href*="imagekit"], a[href*="s3.amazonaws"]'
        );
        var items = [];
        els.forEach(function(el) {
            items.push({
                tag: el.tagName,
                href: el.href || '',
                text: el.textContent.trim().substring(0, 100),
                classes: el.className.toString().substring(0, 120)
            });
        });
        return JSON.stringify(items, null, 2);
    })()
    """)
    print(result or "(no results)")

    print("\n--- ALL LINKS ---")
    result = chrome_js("""
    (function() {
        var links = document.querySelectorAll('a[href]');
        var items = [];
        links.forEach(function(el) {
            var rect = el.getBoundingClientRect();
            if (rect.width > 0 && rect.height > 0) {
                items.push({
                    href: el.href,
                    text: el.textContent.trim().substring(0, 100),
                    classes: el.className.toString().substring(0, 80)
                });
            }
        });
        return JSON.stringify(items, null, 2);
    })()
    """)
    print(result or "(no results)")

    print("\n--- ALL IMAGES ---")
    result = chrome_js("""
    (function() {
        var imgs = document.querySelectorAll('img[src]');
        var items = [];
        imgs.forEach(function(el) {
            var rect = el.getBoundingClientRect();
            if (rect.width > 20 && rect.height > 20) {
                items.push({
                    src: el.src.substring(0, 250),
                    alt: el.alt,
                    size: Math.round(rect.width) + 'x' + Math.round(rect.height)
                });
            }
        });
        return JSON.stringify(items, null, 2);
    })()
    """)
    print(result or "(no results)")

    print("\n--- IFRAMES (videos) ---")
    result = chrome_js("""
    (function() {
        var iframes = document.querySelectorAll('iframe');
        var items = [];
        iframes.forEach(function(el) {
            items.push({
                src: el.src,
                width: el.width,
                height: el.height,
                classes: el.className.toString()
            });
        });
        var videos = document.querySelectorAll('video');
        videos.forEach(function(el) {
            items.push({
                tag: 'VIDEO',
                src: el.src || (el.querySelector('source') || {}).src || '',
                poster: el.poster || ''
            });
        });
        return JSON.stringify(items, null, 2);
    })()
    """)
    print(result or "(no results)")

    print("\n--- OUTER HTML (first 10000 chars) ---")
    result = chrome_js("""
    (function() {
        var main = document.querySelector('main .flex-1') || document.querySelector('main') || document.body;
        return main.outerHTML.substring(0, 10000);
    })()
    """)
    print(result or "(no results)")

    print("\n" + "=" * 70)
    print("DETAIL DISCOVERY COMPLETE")
    print("=" * 70)


def discover_dom():
    """Phase 0: Print DOM structure for the plan page."""
    print("\n" + "=" * 70)
    print("DOM DISCOVERY MODE — PLAN PAGE")
    print("=" * 70)

    url = get_current_url()
    print(f"Current URL: {url}")

    # Focused discovery now that we know the structure
    print("\n--- WEEK NAVIGATION ---")
    result = chrome_js("""
    (function() {
        var prev = document.querySelector('button[title="Previous week"]');
        var next = document.querySelector('button[title="Next week"]');
        var header = document.querySelector('span.text-lg.font-semibold');
        return JSON.stringify({
            prev_found: !!prev,
            prev_disabled: prev ? prev.disabled : null,
            next_found: !!next,
            next_disabled: next ? next.disabled : null,
            week_header: header ? header.textContent.trim() : null
        }, null, 2);
    })()
    """)
    print(result or "(no results)")

    print("\n--- CAMPAIGN CARDS ---")
    result = chrome_js("""
    (function() {
        var cards = document.querySelectorAll('a.group.block');
        var items = [];
        cards.forEach(function(el) {
            items.push({
                href: el.href,
                text: el.textContent.trim().substring(0, 200),
                rect: {
                    x: Math.round(el.getBoundingClientRect().x),
                    y: Math.round(el.getBoundingClientRect().y)
                }
            });
        });
        return JSON.stringify(items, null, 2);
    })()
    """)
    print(result or "(no results)")

    print("\n--- DAY STRUCTURE ---")
    result = chrome_js("""
    (function() {
        // The plan page shows days with their campaigns
        var body = document.querySelector('main') || document.body;
        var text = body.innerText;
        // Extract just the plan content section
        var planStart = text.indexOf('Mar');
        if (planStart === -1) planStart = 0;
        return text.substring(planStart, planStart + 3000);
    })()
    """)
    print(result or "(no results)")

    print("\nDISCOVERY COMPLETE — selectors verified.")


# ─── Plan Page Scraping ──────────────────────────────────────────────────────

def get_week_header():
    """Get the current week's date range text (e.g. 'Mar 16 - 20')."""
    result = chrome_js("""
    (function() {
        var header = document.querySelector('span.text-lg.font-semibold.text-foreground');
        if (!header) header = document.querySelector('span.text-lg.font-semibold');
        return header ? header.textContent.trim() : '';
    })()
    """)
    return result or ""


def parse_week_date(header_text, reference_year=2026):
    """
    Parse week header like 'Mar 16 - 20' into a datetime for the Monday.
    The header doesn't include the year, so we infer it.
    We start at Mar 16 2026 (current) and go backwards, so we track the year.
    """
    if not header_text:
        return None

    # Pattern: "Mar 16 - 20" or "Feb 24 - 28" or "Dec 29 - Jan 2"
    m = re.match(r'(\w+)\s+(\d+)\s*[-–]\s*(?:(\w+)\s+)?(\d+)', header_text)
    if m:
        month_str = m.group(1)
        day = int(m.group(2))
        # Try to parse month
        for fmt in ["%B", "%b"]:
            try:
                month = datetime.strptime(month_str, fmt).month
                return datetime(reference_year, month, day)
            except ValueError:
                continue
    return None


def scrape_plan_page():
    """
    Scrape the current week's plan page to get all campaign cards organized by day.
    Returns: {
        "header": "Mar 16 - 20",
        "days": [
            {"day_name": "Monday", "date": "Mar 16", "campaigns": [
                {"category": "Email Campaigns", "title": "Deal of the Week",
                 "url": "https://..."}
            ]}
        ]
    }
    """
    result = chrome_js("""
    (function() {
        var data = {days: [], header: '', all_text: ''};

        // Get week header
        var headerEl = document.querySelector('span.text-lg.font-semibold.text-foreground');
        if (!headerEl) headerEl = document.querySelector('span.text-lg.font-semibold');
        data.header = headerEl ? headerEl.textContent.trim() : '';

        // Get the main content area (inside the plan page, not the sidebar)
        var mainContent = document.querySelector('.flex-1.overflow-hidden') ||
                          document.querySelector('main.flex-1') ||
                          document.querySelector('main');
        if (!mainContent) {
            data.error = 'Could not find main content';
            return JSON.stringify(data);
        }

        data.all_text = mainContent.innerText.substring(0, 5000);

        // Find all campaign card links
        var cards = mainContent.querySelectorAll('a.group.block[href*="/campaigns/"]');
        var cardData = [];
        cards.forEach(function(card) {
            var rect = card.getBoundingClientRect();
            cardData.push({
                href: card.href,
                text: card.textContent.trim(),
                y: rect.y
            });
        });
        data.cards = cardData;

        return JSON.stringify(data);
    })()
    """)

    if not result:
        return None

    try:
        raw = json.loads(result)
    except json.JSONDecodeError:
        print(f"  WARNING: Could not parse plan page data")
        return None

    # Parse the text to associate cards with days
    all_text = raw.get("all_text", "")
    cards = raw.get("cards", [])
    header = raw.get("header", "")

    # Parse days from the text
    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    parsed_days = []

    lines = all_text.split('\n')
    current_day = None
    current_date = None

    for i, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue

        # Check for day name
        if line in days_of_week:
            if current_day:
                parsed_days.append({"day_name": current_day, "date": current_date, "campaigns": []})
            current_day = line
            # Next line is usually the date (e.g., "Mar 16")
            if i + 1 < len(lines):
                current_date = lines[i + 1].strip()
            continue

    # Don't forget the last day
    if current_day:
        parsed_days.append({"day_name": current_day, "date": current_date, "campaigns": []})

    # Now match campaign cards to days based on their text content
    # Each card's text contains the category and title, e.g. "Email CampaignsDeal of the Week"
    for card in cards:
        card_text = card.get("text", "")
        card_href = card.get("href", "")

        # Extract category and title from the card text
        # The category is the first part, title follows
        category = ""
        title = card_text
        for cat_name in ["Email Campaigns", "Social Shareables", "Phone & Text Scripts",
                         "Direct Mail Templates", "Voice Call & Text Scripts"]:
            if card_text.startswith(cat_name):
                category = cat_name
                title = card_text[len(cat_name):].strip()
                break

        campaign = {
            "category": category,
            "title": title,
            "url": card_href,
            "slug": card_href.split("/")[-1] if card_href else "",
        }

        # Try to match this card to a day by checking if the title appears
        # in the text after the day header
        matched = False
        for day in parsed_days:
            # Simple matching: check if the campaign title appears in the plan text
            # after this day's position
            day_idx = all_text.find(day["day_name"])
            if day_idx >= 0:
                # Find next day's position
                next_day_idx = len(all_text)
                for d in days_of_week:
                    next_pos = all_text.find(d, day_idx + len(day["day_name"]))
                    if next_pos > day_idx and next_pos < next_day_idx:
                        next_day_idx = next_pos

                day_section = all_text[day_idx:next_day_idx]
                if title and title in day_section:
                    day["campaigns"].append(campaign)
                    matched = True
                    break

        if not matched and parsed_days:
            # Fallback: add to last day
            parsed_days[-1]["campaigns"].append(campaign)

    return {
        "header": header,
        "days": parsed_days,
        "raw_cards": cards,
    }


def click_previous_week():
    """Click the previous week navigation arrow."""
    result = chrome_js("""
    (function() {
        var btn = document.querySelector('button[title="Previous week"]');
        if (btn && !btn.disabled) {
            btn.click();
            return 'CLICKED';
        }
        if (btn && btn.disabled) return 'DISABLED';
        return 'NOT_FOUND';
    })()
    """)
    if result == "CLICKED":
        return True
    elif result == "DISABLED":
        print("  Previous week button is disabled (no more weeks)")
        return False
    else:
        print("  WARNING: Could not find previous week button")
        return False


def wait_for_week_change(old_header, timeout=15):
    """Wait for the week header text to change after clicking prev."""
    for _ in range(timeout * 2):
        new_header = get_week_header()
        if new_header and new_header != old_header:
            return True
        time.sleep(0.5)
    return False


# ─── Campaign Detail Page Scraping ──────────────────────────────────────────

def scrape_campaign_detail(url):
    """
    Navigate to a campaign detail page and scrape all content.
    The page has <section> elements for: Introduction, Recommended Audience,
    How to Execute (with Copy blocks for Subject, Preview Text, Body, Scripts),
    and Examples.
    Returns structured dict with sections, copy blocks, images, videos.
    """
    navigate_to(url)
    time.sleep(3)

    # Wait for content to load
    wait_for_element('section', timeout=10)
    time.sleep(1)

    # First: scrape structured sections
    result = chrome_js("""
    (function() {
        var data = {sections: [], copy_blocks: [], images: [], videos: [], links: []};

        // Find the content container (skip sidebar)
        var container = document.querySelector('.w-full.h-full.bg-background') ||
                        document.querySelector('main .flex-1') ||
                        document.querySelector('main') || document.body;

        // Get full text (excluding sidebar)
        data.full_text = container.innerText.substring(0, 20000);

        // Extract structured sections
        var sections = container.querySelectorAll('section');
        sections.forEach(function(sec) {
            var heading = sec.querySelector('h2, h3, h4, [class*="heading"]');
            var headingText = heading ? heading.textContent.trim() : '';
            data.sections.push({
                heading: headingText,
                text: sec.innerText.substring(0, 5000),
                html: sec.innerHTML.substring(0, 8000)
            });
        });

        // Extract "Copy" button content blocks
        // Strategy: Find each Copy button, then look at its direct container
        // to get the label + copyable text
        var allButtons = container.querySelectorAll('button');
        var copyBtnIndex = 0;
        allButtons.forEach(function(btn) {
            var btnText = btn.textContent.trim();
            if (btnText !== 'Copy' && btnText !== 'Copied!') return;

            copyBtnIndex++;
            // The Copy button sits inside a container with the copyable content
            // Walk up to find a container that has both the button and text content
            var parent = btn.parentElement;
            for (var i = 0; i < 6 && parent; i++) {
                // Skip if parent is a section (too broad)
                if (parent.tagName === 'SECTION') break;

                var children = parent.children;
                var contentParts = [];
                var foundBtn = false;

                for (var j = 0; j < children.length; j++) {
                    var child = children[j];
                    // Check if this child contains our button
                    if (child === btn || child.contains(btn)) {
                        foundBtn = true;
                        continue;
                    }
                    var t = child.textContent.trim();
                    // Skip other buttons and "Copy" labels
                    if (child.tagName === 'BUTTON') continue;
                    if (t === 'Copy' || t === 'Copied!') continue;
                    if (t.length > 3) contentParts.push(t);
                }

                if (foundBtn && contentParts.length > 0) {
                    var content = contentParts.join('\\n');

                    // Now find the label — look at the previous sibling of this container
                    var label = '';
                    var prevSib = parent.previousElementSibling;
                    while (prevSib) {
                        var prevText = prevSib.textContent.trim();
                        // Labels are short text like "Subject", "Body", "Text Message 1"
                        if (prevText.length > 0 && prevText.length < 60 && prevText !== 'Copy') {
                            label = prevText;
                            break;
                        }
                        prevSib = prevSib.previousElementSibling;
                    }
                    if (!label) label = 'Copy Block ' + copyBtnIndex;

                    data.copy_blocks.push({
                        label: label,
                        content: content.substring(0, 3000)
                    });
                    break;
                }
                parent = parent.parentElement;
            }
        });

        // Get content images (skip logo, skip tiny icons)
        container.querySelectorAll('img[src]').forEach(function(img) {
            var rect = img.getBoundingClientRect();
            if (rect.width > 80 && rect.height > 80) {
                var src = img.src;
                // For Next.js optimized images, extract the original URL
                if (src.indexOf('/_next/image') > -1) {
                    try {
                        var urlObj = new URL(src);
                        var origUrl = urlObj.searchParams.get('url');
                        if (origUrl) src = origUrl;
                    } catch(e) {}
                }
                if (src.indexOf('logo') === -1) {
                    data.images.push({
                        src: src,
                        alt: img.alt || '',
                        size: Math.round(rect.width) + 'x' + Math.round(rect.height)
                    });
                }
            }
        });

        // Get iframes and videos
        container.querySelectorAll('iframe').forEach(function(f) {
            if (f.src) data.videos.push({src: f.src, type: 'iframe'});
        });
        container.querySelectorAll('video').forEach(function(v) {
            var src = v.src || '';
            var source = v.querySelector('source');
            if (source && source.src) src = source.src;
            if (src) data.videos.push({src: src, type: 'video'});
        });

        // Get content links (external tools, resources)
        container.querySelectorAll('a[href]').forEach(function(a) {
            var rect = a.getBoundingClientRect();
            var href = a.href || '';
            if (rect.width > 0 && rect.height > 0 && href.indexOf('listingleads.com/campaigns') === -1
                && href.indexOf('listingleads.com/plan') === -1
                && href.indexOf('listingleads.com/social') === -1
                && href.indexOf('listingleads.com/favorites') === -1
                && href.indexOf('listingleads.com/blueprints') === -1
                && href.indexOf('listingleads.com/coaching') === -1
                && href.indexOf('listingleads.com/viral') === -1) {
                data.links.push({
                    href: href,
                    text: a.textContent.trim().substring(0, 150),
                    download: a.getAttribute('download') || ''
                });
            }
        });

        // Look for download buttons
        data.download_buttons = [];
        container.querySelectorAll('a[download], a[href*="download"], a[href*=".pdf"], a[href*=".zip"], a[href*=".docx"]').forEach(function(el) {
            data.download_buttons.push({
                href: el.href || '',
                text: el.textContent.trim().substring(0, 100)
            });
        });

        return JSON.stringify(data);
    })()
    """)

    if not result:
        return None

    try:
        return json.loads(result)
    except json.JSONDecodeError:
        return {"full_text": result}


# ─── Cookie Extraction & File Downloads ──────────────────────────────────────

def get_chrome_cookies():
    """Extract cookies from Chrome for authenticated downloads."""
    cookies_str = chrome_js("document.cookie")
    jar = {}
    if cookies_str:
        for pair in cookies_str.split(';'):
            pair = pair.strip()
            if '=' in pair:
                k, v = pair.split('=', 1)
                jar[k.strip()] = v.strip()
    return jar


def download_file(url, dest_path, cookies=None):
    """Download a file from a URL to the destination path."""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        }
        resp = requests.get(url, headers=headers, timeout=30, stream=True, allow_redirects=True)
        if resp.status_code in (401, 403) and cookies:
            resp = requests.get(url, headers=headers, cookies=cookies, timeout=30,
                                stream=True, allow_redirects=True)

        if resp.status_code == 200:
            dest_path.parent.mkdir(parents=True, exist_ok=True)
            with open(dest_path, 'wb') as f:
                for chunk in resp.iter_content(chunk_size=8192):
                    f.write(chunk)
            size_kb = dest_path.stat().st_size / 1024
            print(f"    Downloaded ({size_kb:.0f} KB): {dest_path.name}")
            return True
        else:
            print(f"    Download failed ({resp.status_code}): {url[:80]}")
            return False
    except Exception as e:
        print(f"    Download error: {e}")
        return False


def guess_extension(url):
    """Guess file extension from URL."""
    parsed = urllib.parse.urlparse(url)
    path = parsed.path.lower()
    for ext in ['.pdf', '.png', '.jpg', '.jpeg', '.gif', '.svg', '.mp4',
                '.zip', '.docx', '.xlsx', '.webp']:
        if ext in path:
            return ext
    # Check content in query params (Next.js image URLs)
    if '_next/image' in url:
        # Next.js image optimization — extract original extension
        query = urllib.parse.parse_qs(parsed.query)
        orig_url = query.get('url', [''])[0]
        if orig_url:
            return guess_extension(orig_url)
        return '.png'
    return ''


# ─── Content Organization ────────────────────────────────────────────────────

def get_category_color(category):
    """Get the Excel fill color for a category."""
    for cat_key, color in CATEGORY_COLORS.items():
        if cat_key.lower() in category.lower():
            return color
    return "F5F5F5"


def organize_campaign(campaign_data, category, title, week_date_str, day_name,
                      download_files=True, cookies=None):
    """
    Save a campaign's content to the appropriate folder structure.
    Returns the paths of saved files.
    """
    saved_files = []

    # Determine folder path from URL category slug
    folder_name = CATEGORY_FOLDERS.get(category, category or "Other")
    week_folder = f"{week_date_str}_Week"
    base_dir = OUTPUT_DIR / folder_name / week_folder
    base_dir.mkdir(parents=True, exist_ok=True)

    safe_title = sanitize_filename(title or "untitled")
    prefix = f"{day_name}_{safe_title}" if day_name else safe_title

    # ─── Save structured content ───

    # 1. Save copy blocks (the most valuable content — templates, scripts, etc.)
    copy_blocks = campaign_data.get("copy_blocks", [])
    if copy_blocks:
        copy_text = ""
        for block in copy_blocks:
            label = block.get("label", "")
            content = block.get("content", "")
            if label:
                copy_text += f"=== {label} ===\n"
            copy_text += f"{content}\n\n"
        if copy_text.strip():
            copy_file = base_dir / f"{prefix}_templates.txt"
            copy_file.write_text(copy_text.strip(), encoding="utf-8")
            saved_files.append(str(copy_file))

    # 2. Save full page content (includes instructions, audience, all text)
    full_text = campaign_data.get("full_text", "")
    if full_text:
        # Clean out sidebar/nav text that leaks into the capture
        # Remove everything before "Back" or the category name
        for marker in ["Back\n", f"{category}\n"]:
            idx = full_text.find(marker)
            if idx >= 0:
                full_text = full_text[idx:]
                break

        text_file = base_dir / f"{prefix}_full.txt"
        text_file.write_text(full_text.strip(), encoding="utf-8")
        saved_files.append(str(text_file))

    # 3. Save sections individually for easy reference
    sections = campaign_data.get("sections", [])
    for sec in sections:
        heading = sec.get("heading", "")
        text = sec.get("text", "")
        if heading and text and heading != "Step Content":
            sec_name = sanitize_filename(heading)
            sec_file = base_dir / f"{prefix}_{sec_name}.txt"
            sec_file.write_text(text.strip(), encoding="utf-8")
            saved_files.append(str(sec_file))

    # 4. Save external resource links
    ext_links = campaign_data.get("links", [])
    if ext_links:
        links_text = "External Resources & Tools\n" + "=" * 30 + "\n\n"
        for link in ext_links:
            href = link.get("href", "")
            text = link.get("text", "")
            if href and "listingleads.com" not in href:
                links_text += f"- {text}\n  {href}\n\n"
        if links_text.count("\n") > 4:  # Has actual links
            links_file = base_dir / f"{prefix}_resources.txt"
            links_file.write_text(links_text.strip(), encoding="utf-8")
            saved_files.append(str(links_file))

    if not download_files:
        return saved_files

    # ─── Download assets ───

    seen_urls = set()

    # Download content images (campaign previews, examples)
    for img in campaign_data.get("images", []):
        src = img.get("src", "")
        if not src or src in seen_urls:
            continue
        seen_urls.add(src)

        ext = guess_extension(src) or '.png'
        alt = sanitize_filename(img.get("alt", "")) or f"image_{len(seen_urls)}"
        dest = base_dir / f"{prefix}_{alt}{ext}"
        if not dest.exists():
            if download_file(src, dest, cookies):
                saved_files.append(str(dest))
            time.sleep(0.5)

    # Download any explicit download links (PDFs, etc.)
    for dl in campaign_data.get("download_buttons", []):
        href = dl.get("href", "")
        if not href or href in seen_urls:
            continue
        seen_urls.add(href)
        ext = guess_extension(href) or '.pdf'
        dl_name = sanitize_filename(dl.get("text", "")) or "download"
        dest = base_dir / f"{prefix}_{dl_name}{ext}"
        if not dest.exists():
            if download_file(href, dest, cookies):
                saved_files.append(str(dest))
            time.sleep(0.5)

    # Save video URLs
    videos = campaign_data.get("videos", [])
    video_links = [v.get("src", "") for v in videos if v.get("src")]
    # Also check links for video URLs
    for link in ext_links:
        href = link.get("href", "")
        if any(v in href.lower() for v in ['youtube.com', 'youtu.be', 'vimeo.com', 'wistia', 'loom']):
            video_links.append(href)

    if video_links:
        vid_file = base_dir / f"{prefix}_videos.txt"
        vid_file.write_text("\n".join(sorted(set(video_links))), encoding="utf-8")
        saved_files.append(str(vid_file))

    return saved_files


# ─── Checkpoint / Resume ─────────────────────────────────────────────────────

def save_checkpoint(data):
    """Save progress to checkpoint file."""
    with open(CHECKPOINT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def load_checkpoint():
    """Load progress from checkpoint file."""
    if CHECKPOINT_FILE.exists():
        with open(CHECKPOINT_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


# ─── Excel Generation ────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
THIN_BORDER = Border(bottom=Side(style="thin", color="DDDDDD"))


def create_excel(all_weeks_data):
    """Generate the Excel workbook modeled after Coffee & Contracts calendar."""
    print("\n--- Generating Excel workbook ---")
    wb = openpyxl.Workbook()

    # Flatten all items
    all_items = []
    for week in all_weeks_data:
        week_date_str = week.get("week_date", "")
        week_dt = None
        try:
            week_dt = datetime.strptime(week_date_str, "%Y-%m-%d")
        except (ValueError, TypeError):
            pass

        for day in week.get("days", []):
            day_name = day.get("day_name", "")
            day_date_str = day.get("date", "")

            # Calculate actual date
            day_offsets = {"Monday": 0, "Tuesday": 1, "Wednesday": 2,
                           "Thursday": 3, "Friday": 4, "Saturday": 5, "Sunday": 6}
            offset = day_offsets.get(day_name, 0)
            item_date = week_dt + timedelta(days=offset) if week_dt else None

            for campaign in day.get("campaigns", []):
                all_items.append({
                    "date": item_date,
                    "date_str": item_date.strftime("%m/%d/%Y") if item_date else "",
                    "month_key": item_date.strftime("%B %Y") if item_date else "Unknown",
                    "day_num": item_date.day if item_date else 0,
                    "dow": day_name[:3],  # Mon, Tue, etc.
                    "dow_full": day_name,
                    "category": campaign.get("category", "Other"),
                    "title": campaign.get("title", ""),
                    "url": campaign.get("url", ""),
                    "body_preview": (campaign.get("detail_text", "") or "")[:300],
                    "has_downloads": bool(campaign.get("downloads")),
                    "has_video": bool(campaign.get("videos")),
                    "week_date": week_date_str,
                })

    # Sort chronologically
    all_items.sort(key=lambda x: x.get("date") or datetime.min)

    # Group by month
    month_order = []
    months = {}
    for item in all_items:
        key = item["month_key"]
        if key not in month_order:
            month_order.append(key)
        if key not in months:
            months[key] = []
        months[key].append(item)

    # Remove default sheet
    default_sheet = wb.active

    # ─── Monthly Sheets ───
    for month_name in month_order:
        items = months[month_name]
        ws = wb.create_sheet(title=month_name[:31])

        # Title row
        ws.merge_cells("A1:H1")
        cell = ws["A1"]
        cell.value = f"Listing Leads — {month_name}"
        cell.font = TITLE_FONT
        cell.fill = TITLE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 8

        # Headers
        headers = ["Day", "DOW", "Date", "Category", "Campaign Title",
                    "Content Preview", "Downloads", "Videos"]
        col_widths = [6, 6, 12, 24, 40, 55, 10, 10]

        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Data rows
        for ri, item in enumerate(items, 4):
            color = get_category_color(item["category"])
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            row_data = [
                item["day_num"],
                item["dow"],
                item["date_str"],
                item["category"],
                item["title"],
                item["body_preview"],
                "Yes" if item["has_downloads"] else "",
                "Yes" if item["has_video"] else "",
            ]

            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = fill
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal="left", vertical="top",
                                        wrap_text=(ci >= 5))
                c.font = Font(name="Calibri", size=10)

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:H{len(items) + 3}"

    # ─── All Content Sheet ───
    ws_all = wb.create_sheet(title="All Content")
    ws_all.merge_cells("A1:I1")
    cell = ws_all["A1"]
    cell.value = "Listing Leads — All Content"
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_all.row_dimensions[1].height = 30
    ws_all.row_dimensions[2].height = 8

    headers = ["Week Of", "Day", "Date", "Category", "Campaign Title",
               "Content Preview", "Campaign URL", "Downloads", "Videos"]
    col_widths = [12, 10, 12, 24, 40, 55, 50, 10, 10]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws_all.cell(row=3, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws_all.column_dimensions[get_column_letter(ci)].width = w

    for ri, item in enumerate(all_items, 4):
        color = get_category_color(item["category"])
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        row_data = [
            item["week_date"],
            item["dow_full"],
            item["date_str"],
            item["category"],
            item["title"],
            item["body_preview"],
            item["url"],
            "Yes" if item["has_downloads"] else "",
            "Yes" if item["has_video"] else "",
        ]

        for ci, val in enumerate(row_data, 1):
            c = ws_all.cell(row=ri, column=ci, value=val)
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=(ci >= 5))
            c.font = Font(name="Calibri", size=10)

    ws_all.freeze_panes = "A4"
    ws_all.auto_filter.ref = f"A3:I{len(all_items) + 3}"

    # ─── Summary Sheet ───
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum.merge_cells("A1:E1")
    cell = ws_sum["A1"]
    cell.value = "Listing Leads — Content Summary"
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30
    ws_sum.row_dimensions[2].height = 8

    # Category breakdown
    ws_sum.cell(row=3, column=1, value="CONTENT BY CATEGORY").font = Font(bold=True, size=12)
    ws_sum.merge_cells("A3:E3")

    sum_headers = ["Category", "Total Items", "% of Total", "Weeks Present"]
    for ci, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        ws_sum.column_dimensions[get_column_letter(ci)].width = [28, 12, 12, 14][ci - 1]

    # Count
    cat_counts = {}
    cat_weeks = {}
    for item in all_items:
        cat = item["category"]
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
        if cat not in cat_weeks:
            cat_weeks[cat] = set()
        cat_weeks[cat].add(item.get("week_date", ""))

    total = max(len(all_items), 1)
    row = 5
    for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
        color = get_category_color(cat)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for ci, val in enumerate([cat, count, f"{count/total*100:.1f}%",
                                   len(cat_weeks.get(cat, set()))], 1):
            c = ws_sum.cell(row=row, column=ci, value=val)
            c.fill = fill
            c.font = Font(name="Calibri", size=10)
        row += 1

    row += 1
    ws_sum.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws_sum.cell(row=row, column=2, value=total).font = Font(bold=True)

    # Weekly overview
    row += 2
    ws_sum.cell(row=row, column=1, value="WEEKLY OVERVIEW").font = Font(bold=True, size=12)
    ws_sum.merge_cells(f"A{row}:E{row}")
    row += 1

    wk_headers = ["Week Of", "Campaigns", "Categories"]
    for ci, h in enumerate(wk_headers, 1):
        c = ws_sum.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1

    for week in sorted(all_weeks_data, key=lambda w: w.get("week_date", "")):
        campaigns = []
        categories = set()
        for day in week.get("days", []):
            for camp in day.get("campaigns", []):
                campaigns.append(camp.get("title", ""))
                categories.add(camp.get("category", ""))
        ws_sum.cell(row=row, column=1, value=week.get("week_date", ""))
        ws_sum.cell(row=row, column=2, value=len(campaigns))
        ws_sum.cell(row=row, column=3, value=", ".join(sorted(categories)))
        row += 1

    # Remove default sheet
    if default_sheet and default_sheet.title == "Sheet":
        wb.remove(default_sheet)

    # Save
    excel_path = OUTPUT_DIR / "Listing Leads Content Calendar.xlsx"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)
    print(f"  Excel saved: {excel_path}")
    return excel_path


# ─── Main Scraping Loop (Two-Phase) ──────────────────────────────────────────

def navigate_to_week_n(n):
    """
    From the plan page (which always loads at current week),
    click 'Previous week' n times to get to the nth previous week.
    Returns the week header text after navigation, or None on failure.
    """
    if n == 0:
        return get_week_header()

    for i in range(n):
        old_header = get_week_header()
        if not click_previous_week():
            print(f"    Failed clicking prev at step {i+1}/{n}")
            return None
        time.sleep(1.5)
        if not wait_for_week_change(old_header, timeout=10):
            # Try once more with longer wait
            time.sleep(2)
            if get_week_header() == old_header:
                print(f"    Week didn't change at step {i+1}/{n}")
                return None
    return get_week_header()


def run_scraper(max_weeks=None, download_files=True, resume=False):
    """
    Two-phase scraping:
      Phase 1: Stay on plan page, navigate through weeks, collect all campaign URLs
      Phase 2: Visit each campaign detail URL to scrape & download content
    """
    print("\n" + "=" * 70)
    print("LISTING LEADS CONTENT SCRAPER")
    print("=" * 70)

    url = get_current_url()
    print(f"\nChrome tab ({TAB}) URL: {url}")
    if "listingleads.com" not in url:
        print(f"\nERROR: Navigate to {PLAN_URL} in Chrome first!")
        return

    # Navigate to plan page if not already there
    if "/plan" not in url:
        print(f"Navigating to {PLAN_URL}...")
        navigate_to(PLAN_URL)
        time.sleep(3)

    # Load checkpoint if resuming
    all_weeks_data = []
    scraped_urls = set()
    if resume and CHECKPOINT_FILE.exists():
        all_weeks_data = load_checkpoint()
        # Collect URLs already scraped (campaigns that have detail_text)
        for w in all_weeks_data:
            for d in w.get("days", []):
                for c in d.get("campaigns", []):
                    if c.get("detail_text"):
                        scraped_urls.add(c.get("url", ""))
        print(f"\nResuming: {len(all_weeks_data)} weeks, {len(scraped_urls)} campaigns already done")

    # ─── PHASE 1: Collect all week data from plan page ───
    print("\n" + "=" * 70)
    print("PHASE 1: Collecting weekly campaign listings")
    print("=" * 70)

    # We track the year manually since headers don't include it.
    # Current week starts at Mar 16, 2026 (week 0).
    # Each "prev" click goes back one week.
    current_year = 2026
    last_month = None  # Track month to detect year rollover
    weeks_to_collect = []

    # Start fresh on plan page for phase 1
    navigate_to(PLAN_URL)
    time.sleep(3)
    wait_for_element('button[title="Previous week"]', timeout=10)
    time.sleep(1)

    week_index = 0  # 0 = current week
    weeks_collected = 0

    while True:
        if max_weeks and weeks_collected >= max_weeks:
            print(f"\n  Reached --weeks limit ({max_weeks})")
            break

        header_text = get_week_header()
        print(f"\n  Week {weeks_collected + 1}: {header_text}", end="")

        if not header_text:
            print(" — ERROR: no header")
            break

        # Parse the date with year tracking
        week_date = parse_week_date(header_text, reference_year=current_year)
        if week_date:
            # Detect year rollover: if month jumped UP while going backwards,
            # we crossed into the previous year (e.g., Jan → Dec)
            current_month = week_date.month
            if last_month is not None and current_month > last_month:
                # Month increased going backwards = year rollover
                current_year -= 1
                week_date = parse_week_date(header_text, reference_year=current_year)

            last_month = week_date.month if week_date else last_month
            week_date_str = week_date.strftime("%Y-%m-%d")
            print(f" → {week_date_str}", end="")

            # Check target
            if week_date < TARGET_DATE:
                print(f" — past target, stopping")
                break
        else:
            week_date_str = sanitize_filename(header_text)
            print(f" — could not parse date", end="")

        # Check if already in checkpoint data
        already_have = any(w.get("week_date") == week_date_str for w in all_weeks_data)
        if already_have:
            print(f" — already have, skipping")
        else:
            # Scrape this week's campaign cards
            plan_data = scrape_plan_page()
            if plan_data:
                days = plan_data.get("days", [])
                total = sum(len(d.get("campaigns", [])) for d in days)
                print(f" — {total} campaigns")

                weeks_to_collect.append({
                    "week_date": week_date_str,
                    "header_text": header_text,
                    "days": days,
                    "week_index": week_index,
                })
            else:
                print(f" — ERROR scraping")

        weeks_collected += 1

        # Click previous week
        old_header = header_text
        if not click_previous_week():
            print("\n  No more previous weeks available.")
            break

        time.sleep(1.5)
        if not wait_for_week_change(old_header, timeout=10):
            time.sleep(2)
            if get_week_header() == old_header:
                print("\n  Navigation stuck.")
                break

        week_index += 1

    total_campaigns = sum(
        sum(len(d.get("campaigns", [])) for d in w.get("days", []))
        for w in weeks_to_collect
    )
    print(f"\n\n  Phase 1 complete: {len(weeks_to_collect)} new weeks, {total_campaigns} campaigns to scrape")

    if not weeks_to_collect and not all_weeks_data:
        print("  Nothing to scrape!")
        return

    # ─── PHASE 2: Visit each campaign detail page ───
    print("\n" + "=" * 70)
    print("PHASE 2: Scraping campaign detail pages")
    print("=" * 70)

    cookies = get_chrome_cookies() if download_files else {}
    campaign_total = 0
    campaign_done = 0

    # Count total campaigns to process
    for week in weeks_to_collect:
        for day in week.get("days", []):
            for camp in day.get("campaigns", []):
                if camp.get("url") and camp.get("url") not in scraped_urls:
                    campaign_total += 1

    print(f"  {campaign_total} campaign detail pages to visit\n")

    for week in weeks_to_collect:
        week_date_str = week["week_date"]
        header_text = week["header_text"]
        days = week["days"]

        print(f"\n{'='*50}")
        print(f"Week: {header_text} ({week_date_str})")
        print(f"{'='*50}")

        for day in days:
            for campaign in day.get("campaigns", []):
                camp_url = campaign.get("url", "")
                camp_title = campaign.get("title", "")
                camp_category = campaign.get("category", "")

                if not camp_url:
                    continue
                if camp_url in scraped_urls:
                    continue

                campaign_done += 1
                print(f"\n  [{campaign_done}/{campaign_total}] {camp_category}: {camp_title}")

                # Navigate to detail page and scrape
                detail = scrape_campaign_detail(camp_url)
                if detail:
                    # Store preview for Excel
                    copy_preview = ""
                    for cb in detail.get("copy_blocks", []):
                        label = cb.get("label", "")
                        content = cb.get("content", "")
                        if label:
                            copy_preview += f"[{label}] "
                        copy_preview += content[:200] + " | "

                    if copy_preview.strip(" |"):
                        campaign["detail_text"] = copy_preview[:500]
                    else:
                        full = detail.get("full_text", "") or ""
                        for marker in ["Introduction\n", "How to Execute\n", "Recommended Audience\n"]:
                            idx = full.find(marker)
                            if idx >= 0:
                                full = full[idx:]
                                break
                        campaign["detail_text"] = full[:500]

                    campaign["downloads"] = detail.get("download_buttons", [])
                    campaign["videos"] = detail.get("videos", [])
                    campaign["images"] = detail.get("images", [])

                    # Save to folders
                    saved = organize_campaign(
                        detail, camp_category, camp_title,
                        week_date_str, day.get("day_name", ""),
                        download_files=download_files, cookies=cookies
                    )
                    print(f"    Saved {len(saved)} files")
                    scraped_urls.add(camp_url)
                else:
                    print(f"    WARNING: Could not scrape detail page")

                time.sleep(1)

        # Save this week's data to the main list
        all_weeks_data.append(week)
        save_checkpoint(all_weeks_data)
        print(f"\n  Week {week_date_str} saved. Checkpoint updated.")

    # Save raw JSON
    raw_dir = OUTPUT_DIR / "_raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    raw_file = raw_dir / "all_weeks.json"
    with open(raw_file, 'w', encoding='utf-8') as f:
        json.dump(all_weeks_data, f, indent=2, ensure_ascii=False)

    # Summary
    print("\n" + "=" * 70)
    print("SCRAPING COMPLETE")
    print("=" * 70)
    print(f"  Weeks scraped: {len(all_weeks_data)}")
    total_all = sum(
        sum(len(d.get("campaigns", [])) for d in w.get("days", []))
        for w in all_weeks_data
    )
    print(f"  Total campaigns: {total_all}")
    dates = [w["week_date"] for w in all_weeks_data if w.get("week_date")]
    if dates:
        print(f"  Date range: {min(dates)} to {max(dates)}")
    print(f"  Raw JSON: {raw_file}")

    # Generate Excel
    create_excel(all_weeks_data)

    print(f"\n  All content saved to: {OUTPUT_DIR}")
    print("  Done!")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Scrape Listing Leads weekly plan content",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 scrape_listing_leads.py --tab 1                    Full scrape
  python3 scrape_listing_leads.py --tab 1 --discover         Plan page discovery
  python3 scrape_listing_leads.py --tab 1 --discover-detail  Campaign detail discovery
  python3 scrape_listing_leads.py --tab 1 --weeks 1          Test with 1 week
  python3 scrape_listing_leads.py --tab 1 --resume           Resume from checkpoint
  python3 scrape_listing_leads.py --tab 1 --no-download      Skip file downloads
  python3 scrape_listing_leads.py --tab 1 --excel-only       Regenerate Excel from JSON
        """
    )
    parser.add_argument("--tab", type=int, default=1,
                        help="Chrome tab number (default: 1)")
    parser.add_argument("--discover", action="store_true",
                        help="Run plan page DOM discovery")
    parser.add_argument("--discover-detail", action="store_true",
                        help="Run campaign detail page DOM discovery")
    parser.add_argument("--weeks", type=int, default=None,
                        help="Max weeks to scrape (default: all back to Sep 2025)")
    parser.add_argument("--resume", action="store_true",
                        help="Resume from checkpoint")
    parser.add_argument("--no-download", action="store_true",
                        help="Skip file downloads (text only)")
    parser.add_argument("--excel-only", action="store_true",
                        help="Regenerate Excel from existing checkpoint JSON")
    parser.add_argument("--output", type=str, default=None,
                        help="Custom output directory")
    args = parser.parse_args()

    global TAB, OUTPUT_DIR
    TAB = f"tab {args.tab}"

    if args.output:
        OUTPUT_DIR = Path(args.output)

    chrome_activate()

    if args.discover:
        discover_dom()
    elif args.discover_detail:
        discover_detail_page()
    elif args.excel_only:
        if CHECKPOINT_FILE.exists():
            data = load_checkpoint()
            print(f"Loaded {len(data)} weeks from checkpoint")
            create_excel(data)
        else:
            print(f"ERROR: No checkpoint at {CHECKPOINT_FILE}")
    else:
        run_scraper(
            max_weeks=args.weeks,
            download_files=not args.no_download,
            resume=args.resume,
        )


if __name__ == "__main__":
    main()
